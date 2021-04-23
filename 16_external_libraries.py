# *************** requests *****************
print("*************** requests *****************")
import requests

result1 = requests.get("https://httpbin.org/get", params={"my_key": 'value', "my_key2": 'value2'})
result2 = requests.post("https://httpbin.org/post", data={"next": 'anything'}) # 'data' can be tuple, dict, jsonstring
result3 = requests.post("https://httpbin.org/post", files={'test_file': open(r'data/template.html', 'rb')}) # sending files

data = result1.text        # response data
content = result1.content  # binary response data
json_data = result2.json()   # json response
print(json_data)
print(result3.json())
print(result2.encoding)    # encoding type
print(result2.status_code) # status_code

# from requests import session
# using session, we can easily manage our credentials


# *************** Web Scrapping *****************
print("*************** Web Scrapping *****************")

from bs4 import BeautifulSoup # like JQuery

response = requests.get("https://stackoverflow.com/questions")
soup = BeautifulSoup(response.text, 'html.parser')

qs_element = soup.select_one('#mainbar').select_one('.fs-body3')
qs_count = int(qs_element.getText().split()[0].replace(',', ''))
print('Found :', qs_count, 'questions')

all_qs_data = soup.select_one('#questions').select('.question-summary .question-hyperlink')
print('Showing first', len(all_qs_data), 'questions')
for index, qs in enumerate(all_qs_data):
    print('   ', index+1, ' :' if index < 9 else ':', qs.getText().strip())




# # *************** Automating Web Browser *****************
print("*************** Automating Web Browser *****************")

# Download we driver for a browser : https://pypi.org/project/selenium/

from selenium import webdriver


options = webdriver.ChromeOptions()
options.binary_location = r"C:\Users\Mim\AppData\Local\Google\Chrome\Application\chrome.exe"
chrome_driver_binary = r"C:\Users\Nayeem\Downloads\chromedriver.exe"
browser = webdriver.Chrome(chrome_driver_binary, options=options)

browser.get("https://www.facebook.com/")

username_box = browser.find_element_by_id('email')
username_box.send_keys("019123456789")
password_box = browser.find_element_by_id('pass')
password_box.send_keys("password12354")
password_box.submit()

# assert "MD Hasan" in browser.page_source
# browser.quit()




# # # *************** PyPDF2 *****************
print("*************** PyPDF2 *****************")

import PyPDF2

# Reading and Writing Pdf
with open('data\\pdf_file1.pdf', 'rb') as file:
    reader = PyPDF2.PdfFileReader(file)
    print(reader.numPages)
    page1 = reader.getPage(0)
    page1.rotateClockwise(90)
    writer = PyPDF2.PdfFileWriter()
    writer.addPage(page1)
    with open('data\\pdf_file2.pdf', 'wb') as output:
        writer.write(output)

# merging Pdf
merger = PyPDF2.PdfFileMerger()
file_names = ['data\\pdf_file1.pdf', 'data\\pdf_file2.pdf']
for file in file_names:
    merger.append(file)
merger.write('data\\pdf_file3.pdf')




# # # # *************** Working with Exel *****************
print("*************** Working with Exel *****************")

import openpyxl

# wb = openpyxl.Workbook()    # Create an empty workbook
wb = openpyxl.load_workbook("data\\my_workbook.xlsx")  # Load an workbook
print('Sheet Names : ', wb.sheetnames)

sheet1 = wb["Sheet1"]
cell = sheet1["a1"]

print(cell.row)
print(cell.column)
print(cell.coordinate)
print(cell.value)

for row in range(1, sheet1.max_row + 1):
    for column in range(1, sheet1.max_column + 1):
        cell = sheet1.cell(row, column)
        print(cell.value, end=' ')
    print()

# cells = sheet1["b"]
# cells = sheet1["a:c"]
# cells = sheet1["a1:c4"]
# cells = sheet1[1:3]

sheet1.append(["Kaka", 62, 86])
wb.save("data\\my_workbook2.xlsx")




# # # *************** Numpy *****************
print("*************** Numpy *****************")

import numpy as np

array = np.array([2, 5, 9]) # 1D array
print(array)
print(type(array))

array = np.array([[2, 5, 9], [5, 8, 9]]) # 2D array (metrics)

print('Shape', array.shape)

array1 = np.zeros((3, 4), dtype=int) # shape : (3, 4)
array2 = np.ones((3, 4), dtype=int) # shape : (3, 4)
array3 = np.full((3, 4), 5, dtype=int) # shape : (3, 4)
array4 = np.random.random((3, 4)) # shape : (3, 4)
# print(array1)
# print(array3)
# print(array4)
# print(array4.max())
# print(array4.min())
# print(array4.dtype)
# print(array4.T)  # rows to culumn and vice-versa
# print(array3[0, 0])  # accessing element

print(array4 > 0.5)
array5 = array4[array4 > 0.5]
print(array5)

print(np.sum(array5))
print(np.round(array5))
 
print(array2 + array3)
print(array2 + 2)

dimensions_inch = np.array([4, 8, 3])
dimensions_cm = dimensions_inch * 2.54
print(dimensions_inch)
print(dimensions_cm)


